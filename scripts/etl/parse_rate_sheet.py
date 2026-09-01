#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ETL de rate sheets de frete marítimo: desagrupa a planilha do armador e devolve
um modelo relacional normalizado.

O problema que este script resolve
----------------------------------
Uma linha da rate sheet agrupa MUITAS cotações. Exemplo real (Brasil!L14):

    PIL | Xiamen | Navegantes/Itapoa/Paranagua | 8800 | 9700，（9400<=16ton cargo weight) | ...

Isso é 1 armador x 1 POL x 3 PODs x 3 equipamentos = 9 cotações, sendo que uma
delas tem tarifa condicional a peso. A planilha inteira (139 linhas úteis)
esconde alguns milhares de cotações.

Padrões de agrupamento tratados (todos observados no arquivo real)
------------------------------------------------------------------
 1. Carrier em célula MESCLADA verticalmente          -> forward fill
 2. POL/POD múltiplos separados por "/"               -> produto cartesiano
 3. Tarifa condicional a peso, 4 sintaxes diferentes  -> weight_operator/basis
 4. Tarifa com override POR POD: "8910，（RIO, 8610)"  -> tarifa específica do POD
 5. Tarifa restrita a mercadoria: "8500 (Tyre)"       -> cargo_type
 6. Tarifa com adicional embutido: "9000+pct usd100"  -> vira surcharge
 7. Equipamento sobrescrito: "7415 /40reefer"         -> 40RF, não 40NOR
 8. Coluna de equipamento contendo TAXA, não tarifa   -> só surcharge
 9. Portos colados por erro: "Tianjin/XingangQingdao" -> de-glue automático
10. Validade sem ano, em 5 formatos, às vezes inválida (11.31)
11. Free time composto "21 /18 days"
12. Taxas condicionais no Remark: OWS por faixa de peso

Nada é descartado em silêncio: o que o parser não consegue afirmar com certeza
vira uma linha em `rate_issues` — a trilha de auditoria da carga.

Uso:
    python scripts/etl/parse_rate_sheet.py "<caminho.xlsx>" [--year 2026]
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import json
import os
import re
import sys
import unicodedata
from collections import OrderedDict

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# --------------------------------------------------------------------------
# 1. Dicionário de portos (UN/LOCODE)
# --------------------------------------------------------------------------
# Cada entrada: locode -> (nome canônico, país, [apelidos e grafias erradas])
# Os apelidos vêm da planilha real: a mesma cidade aparece como "Ninbgo",
# "Yiantian", "Rio De Janeir", "Port kelang"...
PORTOS = {
    # --- China ---
    # Xingang É o porto novo de Tianjin. A planilha escreve "Tianjin/Xingang",
    # que o split por "/" transformaria em dois portos — são o mesmo.
    "CNTXG": ("Tianjin/Xingang", "CN", ["tianjin", "xingang", "tianjin xingang"]),
    "CNTAO": ("Qingdao", "CN", ["qingdao", "qing dao"]),
    "CNSHA": ("Shanghai", "CN", ["shanghai"]),
    "CNNGB": ("Ningbo", "CN", ["ningbo", "ninbgo"]),
    "CNYTN": ("Yantian", "CN", ["yantian", "yiantian", "yantain"]),
    "CNSHK": ("Shekou", "CN", ["shekou", "shokou"]),
    "CNSZX": ("Shenzhen", "CN", ["shenzhen"]),
    "CNXMN": ("Xiamen", "CN", ["xiamen", "ixiamen"]),
    "CNFOC": ("Fuzhou", "CN", ["fuzhou"]),
    "CNDLC": ("Dalian", "CN", ["dalian"]),
    "CNHUA": ("Huangpu", "CN", ["huangpu"]),
    "CNZSN": ("Zhongshan", "CN", ["zhongshan"]),
    "CNXLN": ("Xiaolan", "CN", ["xiaolan"]),
    "CNZUH": ("Zhuhai", "CN", ["zhuhai"]),
    "CNSHD": ("Shunde", "CN", ["shunde"]),
    "CNJMN": ("Jiangmen", "CN", ["jiangmen"]),
    "CNSNS": ("Sanshui", "CN", ["sanshui"]),
    "CNNSA": ("Nansha", "CN", ["nansha"]),
    "CNJIU": ("Jiujiang", "CN", ["jiujiang"]),
    "CNGMI": ("Gaoming", "CN", ["gaoming"]),
    "CNFOS": ("Foshan", "CN", ["foshan"]),
    "CNZHA": ("Zhanjiang", "CN", ["zhanjiang"]),
    "CNQZH": ("Qinzhou", "CN", ["qinzhou"]),
    "CNWUZ": ("Wuzhou", "CN", ["wuzhou"]),
    "CNSWA": ("Shantou", "CN", ["shantou"]),
    "CNNKG": ("Nanjing", "CN", ["nanjing"]),
    "CNNTG": ("Nantong", "CN", ["nantong"]),
    "CNZJG": ("Zhangjiagang", "CN", ["zhangjiagang"]),
    "CNZHE": ("Zhenjiang", "CN", ["zhenjiang"]),
    "CNYZH": ("Yangzhou", "CN", ["yangzhou"]),
    "CNCZX": ("Changzhou", "CN", ["changzhou", "chagnzhou"]),
    "CNWHI": ("Wuhu", "CN", ["wuhu"]),
    "CNAQG": ("Anqing", "CN", ["anqing"]),
    "CNNCH": ("Nanchang", "CN", ["nanchang"]),
    "CNWUH": ("Wuhan", "CN", ["wuhan"]),
    "CNLYG": ("Lianyungang", "CN", ["lianyungang"]),
    "CNYIC": ("Yichang", "CN", ["yichang"]),
    "CNCKG": ("Chongqing", "CN", ["chongqing"]),
    "CNCNG": ("Changshu", "CN", ["changshu"]),
    "CNJIA": ("Jiangyin", "CN", ["jiangyin"]),
    "CNTAC": ("Taicang", "CN", ["taicang"]),
    "CNTZO": ("Taizhou", "CN", ["taizhou"]),
    "HKHKG": ("Hong Kong", "HK", ["hong kong", "hongkong"]),
    # --- Sudeste asiático ---
    "MYPKG": ("Port Klang", "MY", ["port klang", "port kelang"]),
    "VNSGN": ("Ho Chi Minh", "VN", ["ho chi minh"]),
    "THLCH": ("Laem Chabang", "TH", ["laem chabang"]),
    "THBKK": ("Bangkok", "TH", ["bangkok"]),
    "IDSRG": ("Semarang", "ID", ["semarang"]),
    # --- Brasil ---
    "BRSSZ": ("Santos", "BR", ["santos", "ssz"]),
    "BRRIO": ("Rio de Janeiro", "BR", ["rio de janeiro", "rio de janeir", "rio", "rj"]),
    "BRPNG": ("Paranaguá", "BR", ["paranagua"]),
    "BRNVT": ("Navegantes", "BR", ["navegantes", "navegante"]),
    "BRIOA": ("Itapoá", "BR", ["itapoa"]),
    "BRITJ": ("Itajaí", "BR", ["itajai"]),
    "BRRIG": ("Rio Grande", "BR", ["rio grande"]),
    "BRSUA": ("Suape", "BR", ["suape"]),
    "BRSSA": ("Salvador", "BR", ["salvador"]),
    "BRPEC": ("Pecém", "BR", ["pecem"]),
    "BRVIX": ("Vitória", "BR", ["vitoria"]),
    "BRMAO": ("Manaus", "BR", ["manaus"]),
    "BRVLC": ("Vila do Conde", "BR", ["vila do conde"]),
    # --- Cone Sul ---
    "UYMVD": ("Montevidéu", "UY", ["montevideo"]),
    "ARBUE": ("Buenos Aires", "AR", ["buenos aires", "bue"]),
    "PYASU": ("Assunção", "PY", ["asuncion"]),
}

APELIDO_PARA_LOCODE = {}
for _loc, (_nome, _pais, _apelidos) in PORTOS.items():
    for _a in _apelidos:
        APELIDO_PARA_LOCODE[_a] = _loc


# --------------------------------------------------------------------------
# 2. Normalização de texto
# --------------------------------------------------------------------------
# A planilha foi editada em ambiente CJK: vírgulas, parênteses e til aparecem
# em largura plena (U+FF0C, U+FF08...). Sem isto, nenhum regex casa.
LARGURA_PLENA = {
    "，": ",", "（": "(", "）": ")", "：": ":",
    "～": "~", "、": ",", "。": ".", "；": ";",
    "／": "/", "＋": "+", "＜": "<", "＞": ">",
    "＝": "=", " ": " ",
}


def norm(v) -> str:
    """Texto cru -> ASCII pontuado, espaços colapsados. Preserva o conteúdo."""
    if v is None:
        return ""
    s = str(v)
    for k, r in LARGURA_PLENA.items():
        s = s.replace(k, r)
    s = unicodedata.normalize("NFKC", s)
    return re.sub(r"\s+", " ", s).strip()


def sem_acento(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def chave(s: str) -> str:
    """Chave de busca de porto: minúsculo, sem acento, sem pontuação."""
    s = sem_acento(norm(s)).lower()
    return re.sub(r"[^a-z0-9 ]+", " ", s).strip()


# --------------------------------------------------------------------------
# 3. Resolução de portos
# --------------------------------------------------------------------------
def resolve_porto(token: str):
    """Token -> locode, ou None se desconhecido."""
    return APELIDO_PARA_LOCODE.get(chave(token))


def desgruda(token: str):
    """
    "XingangQingdao" -> ["CNTXG", "CNTAO"].

    A planilha tem barras faltando (Brasil!L53: "Tianjin/XingangQingdao/...").
    Testa cada ponto de corte; só aceita se AMBAS as metades forem portos
    conhecidos — assim não inventa divisões.
    """
    k = chave(token)
    for i in range(3, len(k) - 2):
        a, b = APELIDO_PARA_LOCODE.get(k[:i].strip()), APELIDO_PARA_LOCODE.get(k[i:].strip())
        if a and b:
            return [a, b]
    return None


PAREN = re.compile(r"\(([^)]*)\)")


def split_portos(celula: str):
    """
    Célula de POL/POD -> (locodes, notas, nao_resolvidos).

    Trata: separador "/", barras duplas, anotações entre parênteses
    ("Ningbo(ETD Apr.01)", "Qingdao (Tyre)") e portos colados.
    """
    txt = norm(celula)
    if not txt:
        return [], [], []
    locodes, notas, desconhecidos = [], [], []
    for bruto in txt.split("/"):
        bruto = bruto.strip()
        if not bruto:
            continue  # barra dupla: "Rio De Janeiro //Santos"
        for nota in PAREN.findall(bruto):
            if nota.strip():
                notas.append(nota.strip())
        nome = PAREN.sub("", bruto).strip(" ,.")
        if not nome:
            # A anotação vinha sozinha depois da barra; nada a resolver aqui.
            continue
        loc = resolve_porto(nome)
        if loc:
            locodes.append(loc)
            continue
        partes = desgruda(nome)
        if partes:
            locodes.extend(partes)
            notas.append(f"portos colados na origem: '{nome}'")
            continue
        desconhecidos.append(nome)
    # Dedupe preservando ordem. "Tianjin/Xingang" colapsa em CNTXG uma vez só.
    return list(OrderedDict.fromkeys(locodes)), notas, desconhecidos


# --------------------------------------------------------------------------
# 4. Parsing das células de tarifa
# --------------------------------------------------------------------------
CARGAS = {
    "tyre": "PNEU", "tyres": "PNEU",
    "textile": "TEXTIL", "yarn": "TEXTIL", "fabric": "TEXTIL",
    "solar": "SOLAR", "reefer": "REEFER",
}

# "9400<=16ton cargo weight" | "8410<=10TON VGM" | "8410<VGM 10TON"
COND_PESO = re.compile(
    r"(?P<valor>\d{3,6})?\s*(?P<op><=|<|>=|>)\s*"
    r"(?:(?P<basis1>vgm|cargo\s*weight)\s*)?"
    r"(?P<peso>\d{1,3}(?:\.\d+)?)\s*ton?s?\b\s*"
    r"(?P<basis2>vgm|cargo\s*weight)?",
    re.I,
)
# "+pct usd100" / "+local agent H/C usd50"
ADICIONAL = re.compile(r"\+\s*([^+]+)", re.I)
VALOR_USD = re.compile(r"(?:usd\s*(\d+(?:\.\d+)?)|(\d+(?:\.\d+)?)\s*usd)", re.I)
NUM = re.compile(r"\b(\d{3,6})(?:\.0+)?\b")


def _basis(m) -> str:
    b = (m.group("basis1") or m.group("basis2") or "").lower()
    if "vgm" in b:
        return "VGM"          # peso bruto verificado: INCLUI a tara do contêiner
    if "cargo" in b:
        return "CARGO"        # peso da mercadoria
    return "NAO_ESPECIFICADO"


def parse_tarifa(celula, equip_padrao: str):
    """
    Célula da coluna de equipamento -> estrutura da tarifa.

    Devolve dict com: base, condicoes de peso, overrides por POD, cargo_type,
    equipamento efetivo, adicionais embutidos e avisos.
    Se a célula não contiver tarifa (só taxa), base fica None.
    """
    cru = norm(celula)
    r = {
        "raw": cru, "base": None, "equipamento": equip_padrao, "cargo_type": None,
        "condicoes": [], "overrides_pod": [], "adicionais": [], "avisos": [],
    }
    if not cru:
        return r

    # (7) equipamento sobrescrito: "7415 /40reefer", "8420(reefer )"
    if re.search(r"reefer", cru, re.I):
        r["equipamento"] = "40RF"
    if re.search(r"\bLCL\b|/\s*CBM", cru, re.I):
        r["equipamento"] = "LCL"

    # (5) mercadoria a que a tarifa se restringe
    for nota in PAREN.findall(cru):
        for palavra in re.split(r"[,\s]+", nota.lower()):
            if palavra.strip() in CARGAS:
                r["cargo_type"] = CARGAS[palavra.strip()]
    if re.search(r"/\s*40\s*reefer", cru, re.I):
        r["cargo_type"] = "REEFER"

    # (6) adicional embutido na tarifa: "9000+pct usd100"
    corpo = cru
    for extra in ADICIONAL.findall(cru):
        m = VALOR_USD.search(extra)
        if m:
            r["adicionais"].append({
                "raw": extra.strip(),
                "amount": float(m.group(1) or m.group(2)),
                "label": VALOR_USD.sub("", extra).strip(" +/-,"),
            })
    corpo = ADICIONAL.sub("", corpo)

    # (8) coluna de equipamento com TAXA no lugar de tarifa
    if cru.lstrip().startswith("+"):
        r["avisos"].append("celula_de_equipamento_contem_taxa_nao_tarifa")
        return r

    # LCL: "USD1/CBM" — preço por metro cúbico, não por contêiner
    m_cbm = re.search(r"usd\s*(\d+(?:\.\d+)?)\s*/\s*cbm", cru, re.I)
    if m_cbm:
        r["base"] = float(m_cbm.group(1))
        r["equipamento"] = "LCL"
        r["unidade"] = "CBM"
        return r

    # (3) condições de peso — podem existir com ou sem tarifa-base antes
    for m in COND_PESO.finditer(corpo):
        r["condicoes"].append({
            "valor": float(m.group("valor")) if m.group("valor") else None,
            "operador": m.group("op"),
            "peso_ton": float(m.group("peso")),
            "base_peso": _basis(m),
            "raw": m.group(0).strip(),
        })
    resto = COND_PESO.sub(" ", corpo)

    # (4) override por POD: "（RIO, 8610)" — tarifa diferente para um destino
    for nota in PAREN.findall(resto):
        loc = None
        for pedaco in re.split(r"[,;]", nota):
            pedaco = pedaco.strip()
            if not pedaco:
                continue
            cand = resolve_porto(re.sub(r"\d", "", pedaco).strip())
            if cand:
                loc = cand
            mv = NUM.search(pedaco)
            if loc and mv:
                r["overrides_pod"].append({"pod": loc, "valor": float(mv.group(1)), "raw": nota.strip()})
                loc = None
    fora_paren = PAREN.sub(" ", resto)

    m = NUM.search(fora_paren)
    if m:
        r["base"] = float(m.group(1))
    elif r["condicoes"] and r["condicoes"][0]["valor"] is not None:
        # "8410<=10TON VGM": a tarifa só existe sob a condição.
        r["base"] = r["condicoes"][0]["valor"]

    # Sanidade comercial: o desconto por peso tem de ser MENOR que a base.
    # Brasil!E23 traz "1000,（9800<=18ton vgm）" — 1000 é dígito faltando em 10000.
    for c in r["condicoes"]:
        if c["valor"] is not None and r["base"] is not None:
            if c["operador"] in ("<=", "<") and c["valor"] > r["base"]:
                r["avisos"].append(
                    f"tarifa_base_menor_que_desconto (base={r['base']:.0f}, "
                    f"condicional={c['valor']:.0f}) — provavel digito faltando"
                )
    if r["base"] is not None and r["base"] < 1500 and r["equipamento"] != "LCL":
        r["avisos"].append(f"tarifa_implausivel_para_contêiner ({r['base']:.0f} USD)")
    return r


# --------------------------------------------------------------------------
# 5. Validade, free time, taxas
# --------------------------------------------------------------------------
INTERVALO = re.compile(r"\b(\d{1,2})\.(\d{1,2})\s*[~\-]\s*(\d{1,2})\.(\d{1,2})\b")
ETD_NUM = re.compile(r"etd\s*[:\s]?\s*(\d{1,2})[/.](\d{1,2})", re.I)
MESES = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"])}
ETD_MES = re.compile(r"etd\s*([a-z]{3})[a-z]*\.?\s*(\d{1,2})", re.I)
VESSEL = re.compile(r"\(([^)]{4,})\)")


def _data(ano, mes, dia):
    try:
        return dt.date(ano, mes, dia)
    except ValueError:
        return None


def parse_validade(celula, ano: int):
    """
    Coluna 'Validity' -> (inicio, fim, vessel, avisos).

    5 formatos no arquivo: "9.01~9.07", "4.01-4.04", "ETD:9/07",
    "ETD Sep.08", e texto livre com nome de navio. Nenhum traz ano.
    """
    cru = norm(celula)
    avisos = []
    if not cru:
        return None, None, None, ["sem_validade_declarada"]

    vessel = None
    for cand in VESSEL.findall(cru):
        c = cand.strip()
        if re.search(r"[A-Z]{3,}", c) and re.search(r"\d{3,}", c):
            vessel = c[:120]
            break

    m = INTERVALO.search(cru)
    if m:
        m1, d1, m2, d2 = (int(x) for x in m.groups())
        ini, fim = _data(ano, m1, d1), _data(ano, m2, d2)
        if ini is None:
            avisos.append(f"data_inicial_invalida ({m1}.{d1})")
        if fim is None:
            avisos.append(f"data_final_invalida ({m2}.{d2})")
        return ini, fim, vessel, avisos

    m = ETD_NUM.search(cru)
    if m:
        d = _data(ano, int(m.group(1)), int(m.group(2)))
        return d, d, vessel, (["etd_pontual"] if d else ["data_etd_invalida"])

    m = ETD_MES.search(cru)
    if m and m.group(1).lower() in MESES:
        d = _data(ano, MESES[m.group(1).lower()], int(m.group(2)))
        return d, d, vessel, (["etd_pontual"] if d else ["data_etd_invalida"])

    # "11.31" solto (LCL): 31 de novembro não existe.
    m = re.search(r"\b(\d{1,2})\.(\d{1,2})\b", cru)
    if m:
        d = _data(ano, int(m.group(1)), int(m.group(2)))
        if d is None:
            return None, None, vessel, [f"data_inexistente_no_calendario ({m.group(0)})"]
        return d, d, vessel, ["validade_de_dia_unico"]

    return None, None, vessel, ["validade_nao_interpretavel"]


FREE_TIME = re.compile(r"(\d{1,3})\s*(?:/\s*(\d{1,3}))?\s*days?", re.I)


def parse_free_time(celula):
    """'21 /18 days' -> (21, 18). '21 days' -> (21, None)."""
    m = FREE_TIME.search(norm(celula))
    if not m:
        return None, None, norm(celula)
    a = int(m.group(1))
    b = int(m.group(2)) if m.group(2) else None
    return a, b, norm(celula)


CODIGOS_TAXA = ["isps", "css", "spg", "lws", "pct", "ows", "thc", "baf", "caf", "ers"]


def parse_taxas(celula, origem: str):
    """
    Coluna 'Subject to' / 'Remark' -> lista de surcharges.

    Formatos: "isps usd15", "usd30+isps usd14", "css usd15+ spg usd50",
    "MANAUS lws 1400usd from 9/05", "OWS USD200/20GP >14 ton,USD500/20GP >20ton".
    """
    cru = norm(celula)
    if not cru:
        return [], []
    notas, taxas = [], []

    # Sinalizações que NÃO são cobrança: "No OWS", "NO LWS", "Including LWS".
    if re.match(r"^\s*(no|not?)\s+\w+", cru, re.I) or re.search(r"including", cru, re.I):
        return [], [cru]

    ultimo_codigo = None
    for pedaco in re.split(r"[+,]", cru):
        pedaco = pedaco.strip()
        if not pedaco:
            continue
        m = VALOR_USD.search(pedaco)
        if not m:
            if pedaco:
                notas.append(pedaco)
            continue
        valor = float(m.group(1) or m.group(2))
        texto = VALOR_USD.sub(" ", pedaco)

        codigo = next((c.upper() for c in CODIGOS_TAXA if re.search(rf"\b{c}\b", texto, re.I)), None)
        # "OWS USD200/20GP >14 ton, USD500/20GP >20ton": o segundo patamar não
        # repete a sigla. Sem herdá-la, a faixa vira uma taxa órfã.
        if codigo:
            ultimo_codigo = codigo
        elif ultimo_codigo and re.search(r"[<>]=?\s*\d+\s*ton", pedaco, re.I):
            codigo = ultimo_codigo
        equip = None
        me = re.search(r"/\s*(20\s*'?\s*GP|40\s*'?\s*(?:GP|HQ|NOR))", pedaco, re.I)
        if me:
            equip = re.sub(r"[\s']", "", me.group(1)).upper()
        mp = re.search(r"([<>]=?)\s*(\d{1,3})\s*ton", pedaco, re.I)
        peso_min = float(mp.group(2)) if mp else None
        pod = resolve_porto(re.sub(r"[^A-Za-z ]", " ", texto).strip())

        taxas.append({
            "fee_code": codigo or "OTHER",
            "fee_label": re.sub(r"\s+", " ", texto).strip(" ./-") or "taxa não nomeada",
            "amount": valor,
            "currency": "USD",
            "charge_basis": "PER_CONTAINER",
            "equipment_type": equip,
            "min_weight_ton": peso_min,
            "applies_to_pod": pod,
            "condition_raw": pedaco,
            "source_column": origem,
        })
    return taxas, notas


# --------------------------------------------------------------------------
# 6. Leitura da planilha
# --------------------------------------------------------------------------
COLS = {"carrier": 1, "pol": 2, "pod": 3, "20GP": 4, "40HQ": 5, "40NOR": 6,
        "subject_to": 7, "validity": 8, "free_time": 9, "service": 10,
        "remark": 11, "space": 12}
EQUIP_COLS = [("20GP", 4), ("40HQ", 5), ("40NOR", 6)]
# A coluna E cobre "40'GP/40'HQ": uma tarifa só para os dois equipamentos.
EQUIVALENTES = {"40HQ": ["40GP"]}


def cel(ws, linha, col):
    return ws.cell(row=linha, column=col).value


# O mesmo armador aparece com códigos diferentes entre as abas ("MSK" em
# Brasil, "Maersk" na aba LCL). Sem canonizar, o comparador mostra dois.
ALIAS_CARRIER = {
    "maersk": "MSK", "msk": "MSK", "cma cgm": "CMA", "cma": "CMA",
    "evergreen": "EMC", "emc": "EMC", "hapag": "HPL", "hpl": "HPL",
    "yang ming": "YML", "yml": "YML", "cosco": "COSCO", "oocl": "OOCL",
    "pil": "PIL", "hmm": "HMM", "one": "ONE", "msc": "MSC", "cssc": "CSSC",
}


def carrier_scope(bruto: str):
    """'HMM (NAC for Tyre)' -> ('HMM', ['NAC', 'PNEU'])."""
    txt = norm(bruto)
    escopo = []
    for nota in PAREN.findall(txt):
        n = nota.lower()
        if "nac" in n:
            escopo.append("NAC")   # Named Account Contract: tarifa de conta nomeada
        for p, c in CARGAS.items():
            if p in n:
                escopo.append(c)
    nome = PAREN.sub("", txt).strip()
    return ALIAS_CARRIER.get(chave(nome), nome.upper()), list(OrderedDict.fromkeys(escopo))


def processar(caminho: str, ano: int):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    saida = {
        "rate_sheet": {
            "source_file": os.path.basename(caminho).replace(" ", " "),
            "issued_on": None, "currency": "USD",
            "imported_at": dt.datetime.now().isoformat(timespec="seconds"),
            "assumed_year": ano,
        },
        "ports": [], "carriers": [], "routes": [], "quotes": [], "issues": [],
    }
    # A validade mais antiga da planilha não serve como emissão; o nome do
    # arquivo ("0901") é a única pista confiável da data de emissão.
    m = re.search(r"(\d{2})(\d{2})", saida["rate_sheet"]["source_file"])
    if m:
        d = _data(ano, int(m.group(1)), int(m.group(2)))
        saida["rate_sheet"]["issued_on"] = d.isoformat() if d else None

    usados, carriers, rota_id, quote_id = set(), OrderedDict(), 0, 0

    def issue(aba, linha, sev, tipo, detalhe):
        saida["issues"].append({"sheet": aba, "source_row": linha, "severity": sev,
                                "kind": tipo, "detail": detalhe})

    for ws in wb.worksheets:
        aba = ws.title.strip()
        lane = "LCL" if aba.upper().startswith("LCL") else aba
        carrier_atual, escopo_atual = None, []

        for linha in range(2, ws.max_row + 1):
            bruto_carrier = cel(ws, linha, COLS["carrier"])
            if norm(bruto_carrier):
                # (1) célula mesclada: só a primeira linha do bloco traz o nome
                carrier_atual, escopo_atual = carrier_scope(bruto_carrier)
            pol_cel, pod_cel = cel(ws, linha, COLS["pol"]), cel(ws, linha, COLS["pod"])
            if not norm(pol_cel) and not norm(pod_cel):
                continue
            if not carrier_atual:
                issue(aba, linha, "erro", "carrier_ausente", "linha sem armador identificável")
                continue

            pols, notas_pol, desc_pol = split_portos(pol_cel)
            pods, notas_pod, desc_pod = split_portos(pod_cel)
            for d in desc_pol + desc_pod:
                issue(aba, linha, "aviso", "porto_nao_mapeado",
                      f"'{d}' não está no dicionário UN/LOCODE — rota ignorada para esse porto")
            if not pols or not pods:
                issue(aba, linha, "erro", "rota_sem_par",
                      f"POL='{norm(pol_cel)}' POD='{norm(pod_cel)}' — nenhuma rota gerada")
                continue

            ini, fim, vessel, av_val = parse_validade(cel(ws, linha, COLS["validity"]), ano)
            for a in av_val:
                if a.startswith(("data_", "validade_nao")):
                    issue(aba, linha, "aviso", "validade", a)
            ft_pol, ft_pod, ft_raw = parse_free_time(cel(ws, linha, COLS["free_time"]))
            if ft_pol is None and norm(cel(ws, linha, COLS["free_time"])):
                issue(aba, linha, "aviso", "free_time", f"não interpretado: '{ft_raw}'")

            taxas_g, notas_g = parse_taxas(cel(ws, linha, COLS["subject_to"]), "subject_to")
            taxas_k, notas_k = parse_taxas(cel(ws, linha, COLS["remark"]), "remark")
            servico = norm(cel(ws, linha, COLS["service"]))
            espaco = norm(cel(ws, linha, COLS["space"])) or None

            # Equipamentos desta linha
            tarifas = []
            for equip, col in EQUIP_COLS:
                cel_eq = cel(ws, linha, col)
                if not norm(cel_eq):
                    continue
                t = parse_tarifa(cel_eq, equip)
                for a in t["avisos"]:
                    issue(aba, linha, "aviso", "tarifa", f"[{equip}] {a} :: '{t['raw']}'")
                # (8) coluna de equipamento que na verdade traz taxa
                if t["base"] is None and t["adicionais"]:
                    for ad in t["adicionais"]:
                        taxas_g.append({
                            "fee_code": "OTHER", "fee_label": ad["label"] or "adicional",
                            "amount": ad["amount"], "currency": "USD",
                            "charge_basis": "PER_CONTAINER", "equipment_type": None,
                            "min_weight_ton": None, "applies_to_pod": None,
                            "condition_raw": ad["raw"], "source_column": f"col_{equip}",
                        })
                    continue
                if t["base"] is None:
                    continue
                for ad in t["adicionais"]:
                    taxas_g.append({
                        "fee_code": next((c.upper() for c in CODIGOS_TAXA
                                          if re.search(rf"{c}", ad["label"], re.I)), "OTHER"),
                        "fee_label": ad["label"] or "adicional embutido na tarifa",
                        "amount": ad["amount"], "currency": "USD",
                        "charge_basis": "PER_CONTAINER", "equipment_type": equip,
                        "min_weight_ton": None, "applies_to_pod": None,
                        "condition_raw": ad["raw"], "source_column": f"col_{equip}",
                    })
                tarifas.append(t)

            if not tarifas:
                issue(aba, linha, "info", "linha_sem_tarifa",
                      f"{carrier_atual}: sem valor de frete em nenhuma coluna de equipamento")
                continue

            carriers.setdefault(carrier_atual, {"code": carrier_atual, "name": carrier_atual})

            # (2) produto cartesiano POL x POD
            for pol in pols:
                for pod in pods:
                    usados.update((pol, pod))
                    rota_id += 1
                    saida["routes"].append({
                        "id": rota_id, "sheet": aba, "trade_lane": lane, "source_row": linha,
                        "carrier": carrier_atual, "carrier_scope": escopo_atual,
                        "pol": pol, "pod": pod,
                        "service_type": "Transhipment" if "transhipment" in servico.lower()
                                        else ("Direct" if "direct" in servico.lower() else "Outro"),
                        "service_name": servico or None,
                        "validity_start": ini.isoformat() if ini else None,
                        "validity_end": fim.isoformat() if fim else None,
                        "validity_raw": norm(cel(ws, linha, COLS["validity"])) or None,
                        "vessel_ref": vessel,
                        "space_status": espaco,
                        "free_days_pol": ft_pol, "free_days_pod": ft_pod, "free_time_raw": ft_raw,
                        "surcharges": [t for t in (taxas_g + taxas_k)
                                       if t["applies_to_pod"] in (None, pod)],
                        "notes": [n for n in (notas_g + notas_k + notas_pol + notas_pod) if n],
                    })
                    for t in tarifas:
                        # (4) override por POD tem precedência sobre a base
                        ov = next((o for o in t["overrides_pod"] if o["pod"] == pod), None)
                        base = ov["valor"] if ov else t["base"]
                        # Se a linha traz override e este POD não é o do override,
                        # a base vale para os demais destinos — comportamento correto.
                        cond = t["condicoes"][0] if t["condicoes"] else None
                        ajustada = cond["valor"] if (cond and cond["valor"] is not None
                                                     and cond["valor"] != base) else None
                        if ov and ajustada is not None:
                            # Override de POD + condição de peso na mesma célula: o
                            # desconto não é atribuível com segurança ao POD específico.
                            ajustada = None
                        quote_id += 1
                        saida["quotes"].append({
                            "id": quote_id, "route_id": rota_id,
                            "equipment_type": t["equipamento"],
                            "also_valid_for": EQUIVALENTES.get(t["equipamento"], []),
                            "base_rate": base, "currency": "USD",
                            "unit": t.get("unidade", "CONTAINER"),
                            "adjusted_rate": ajustada,
                            "weight_operator": cond["operador"] if cond else None,
                            "weight_limit_ton": cond["peso_ton"] if cond else None,
                            "weight_basis": cond["base_peso"] if cond else None,
                            "cargo_type": t["cargo_type"] or (escopo_atual[-1]
                                          if escopo_atual and escopo_atual[-1] != "NAC" else None),
                            "rate_source": "override_pod" if ov else "base",
                            "raw_cell": t["raw"],
                        })

    saida["carriers"] = list(carriers.values())
    saida["ports"] = [{"unlocode": loc, "name": PORTOS[loc][0], "country": PORTOS[loc][1]}
                      for loc in sorted(usados)]
    return saida


# --------------------------------------------------------------------------
# 7. Emissão
# --------------------------------------------------------------------------
def sql_txt(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"

def _valores(linhas, campos, tabela, chunk=400):
    """
    Gera INSERTs multi-linha na tabela de staging.

    Um INSERT por lote de `chunk` tuplas em vez de um por linha: é o que reduz o
    arquivo de 2 MB para algo que o SQL Editor do Supabase aceita. O boilerplate
    ("insert into ... (col, col, ...) values") passa a ser pago 1x a cada 400
    linhas, não 1x por linha.
    """
    if not linhas:
        return []
    cols = ", ".join(campos)
    saida = []
    for i in range(0, len(linhas), chunk):
        lote = linhas[i:i + chunk]
        tuplas = ",\n".join(
            "(" + ",".join(sql_txt(l.get(c)) for c in campos) + ")" for l in lote
        )
        saida.append(f"insert into mcat.{tabela} ({cols}) values\n{tuplas};")
    return saida


# DDL das tabelas de staging. São tabelas REAIS (não temporárias) porque, no SQL
# Editor, cada arquivo roda numa sessão própria — uma temp table não sobreviveria
# entre as partes. O arquivo final as remove.
DDL_STAGE = """\
-- ============================================================================
-- PARTE A — tabelas de staging
-- Rode esta parte PRIMEIRO. Ela só cria estruturas de apoio; nenhuma tabela
-- definitiva é tocada. Sem tipos nem constraints: staging deve ser tolerante,
-- a validação acontece na parte C.
-- ============================================================================
set search_path to mcat, public;

create table if not exists mcat._stage_route (
  route_key int, carrier text, pol text, pod text, trade_lane text,
  service_type text, service_name text, validity_start text, validity_end text,
  validity_raw text, vessel_ref text, space_status text, sheet text,
  source_row int, carrier_scope text, free_days_pol int, free_days_pod int,
  free_time_raw text
);
create table if not exists mcat._stage_rate (
  route_key int, equipment_type text, also_valid_for text, base_rate numeric,
  currency text, unit text, adjusted_rate numeric, weight_operator text,
  weight_limit_ton numeric, weight_basis text, cargo_type text,
  rate_source text, raw_cell text
);
create table if not exists mcat._stage_surcharge (
  route_key int, fee_code text, fee_label text, amount numeric, currency text,
  charge_basis text, equipment_type text, min_weight_ton numeric,
  condition_raw text, source_column text
);
create table if not exists mcat._stage_issue (
  sheet text, source_row int, severity text, kind text, detail text
);
create table if not exists mcat._stage_map (route_key int, route_id uuid);

truncate mcat._stage_route, mcat._stage_rate, mcat._stage_surcharge,
         mcat._stage_issue, mcat._stage_map;
"""


def _ddl_load(rs, esperado_rotas, esperado_tarifas):
    """PARTE C: transfere do staging para as tabelas definitivas, em conjuntos."""
    return f"""\
-- ============================================================================
-- PARTE C — carga definitiva (rode por ÚLTIMO)
-- Tudo em conjuntos: 6 INSERTs cobrem as ~5.600 linhas. Transação única —
-- qualquer falha desfaz a carga inteira.
-- ============================================================================
set search_path to mcat, public;
begin;

do $$
declare
  v_sheet uuid;
  v_rotas int;
  v_tarifas int;
begin
  insert into mcat.rate_sheets (source_file, issued_on, currency, assumed_year)
  values ({sql_txt(rs['source_file'])}, {sql_txt(rs['issued_on'])}::date,
          {sql_txt(rs['currency'])}, {rs['assumed_year']})
  on conflict (source_file, issued_on) do update set imported_at = now()
  returning id into v_sheet;

  -- Reimportação substitui apenas as linhas DESTA rate sheet (cascade nas filhas).
  delete from mcat.freight_routes where rate_sheet_id = v_sheet;
  delete from mcat.rate_issues    where rate_sheet_id = v_sheet;

  insert into mcat.freight_routes (rate_sheet_id, carrier_id, pol_id, pod_id, trade_lane,
    service_type, service_name, validity_start, validity_end, validity_raw, vessel_ref,
    space_status, source_sheet, source_row, carrier_scope)
  select v_sheet, c.id, o.id, dst.id, t.trade_lane, t.service_type, t.service_name,
         nullif(t.validity_start,'')::date, nullif(t.validity_end,'')::date,
         t.validity_raw, t.vessel_ref, t.space_status, t.sheet, t.source_row, t.carrier_scope
    from mcat._stage_route t
    join mcat.carriers c on c.code     = t.carrier
    join mcat.ports    o on o.unlocode = t.pol
    join mcat.ports  dst on dst.unlocode = t.pod;

  get diagnostics v_rotas = row_count;

  -- Liga o id do ETL ao uuid gravado pela chave natural (aba, linha, POL, POD),
  -- única por construção do ETL — é o que dispensa um INSERT por rota.
  insert into mcat._stage_map (route_key, route_id)
  select t.route_key, r.id
    from mcat.freight_routes r
    join mcat.ports    o on o.id   = r.pol_id
    join mcat.ports  dst on dst.id = r.pod_id
    join mcat._stage_route t
      on t.sheet = r.source_sheet and t.source_row = r.source_row
     and t.pol = o.unlocode and t.pod = dst.unlocode
   where r.rate_sheet_id = v_sheet;

  insert into mcat.equipment_rates (route_id, equipment_type, also_valid_for, base_rate,
    currency, unit, adjusted_rate, weight_operator, weight_limit_ton, weight_basis,
    cargo_type, rate_source, raw_cell)
  select m.route_id, s.equipment_type, s.also_valid_for, s.base_rate, s.currency, s.unit,
         s.adjusted_rate, s.weight_operator, s.weight_limit_ton, s.weight_basis,
         s.cargo_type, s.rate_source, s.raw_cell
    from mcat._stage_rate s join mcat._stage_map m on m.route_key = s.route_key;

  get diagnostics v_tarifas = row_count;

  insert into mcat.rate_surcharges (route_id, fee_code, fee_label, amount, currency,
    charge_basis, equipment_type, min_weight_ton, condition_raw, source_column)
  select m.route_id, s.fee_code, s.fee_label, s.amount, s.currency, s.charge_basis,
         s.equipment_type, s.min_weight_ton, s.condition_raw, s.source_column
    from mcat._stage_surcharge s join mcat._stage_map m on m.route_key = s.route_key;

  insert into mcat.free_time_rules (route_id, free_days_pol, free_days_pod, raw)
  select m.route_id, t.free_days_pol, t.free_days_pod, t.free_time_raw
    from mcat._stage_route t join mcat._stage_map m on m.route_key = t.route_key
   where t.free_days_pol is not null;

  insert into mcat.rate_issues (rate_sheet_id, source_sheet, source_row, severity, kind, detail)
  select v_sheet, sheet, source_row, severity, kind, detail from mcat._stage_issue;

  -- A carga só vale se bater com o que o ETL apurou.
  if v_rotas <> {esperado_rotas} then
    raise exception 'rotas: gravadas %, esperadas {esperado_rotas} (portos ou armadores faltando?)', v_rotas;
  end if;
  if v_tarifas <> {esperado_tarifas} then
    raise exception 'tarifas: gravadas %, esperadas {esperado_tarifas}', v_tarifas;
  end if;
end $$;

drop table if exists mcat._stage_route, mcat._stage_rate, mcat._stage_surcharge,
                     mcat._stage_issue, mcat._stage_map;

commit;

-- Confirmação: deve devolver {esperado_tarifas}.
select count(*) as cotacoes from mcat.v_freight_quotes;
"""


def gerar_partes(d, limite_bytes=350_000):
    """
    Emite a carga em partes que cabem no SQL Editor do Supabase.

    Devolve [(nome_do_arquivo, conteudo)]. As dimensões (portos, armadores) vão
    junto com o staging: são poucas linhas e precisam existir antes da parte C.
    """
    rs = d["rate_sheet"]
    cab = (f"-- GERADO POR scripts/etl/parse_rate_sheet.py — NÃO EDITAR À MÃO.\n"
           f"-- Fonte: {rs['source_file']} | importado em {rs['imported_at']}\n")

    rotas = [{
        "route_key": r["id"], "carrier": r["carrier"], "pol": r["pol"], "pod": r["pod"],
        "trade_lane": r["trade_lane"], "service_type": r["service_type"],
        "service_name": r["service_name"], "validity_start": r["validity_start"],
        "validity_end": r["validity_end"], "validity_raw": r["validity_raw"],
        "vessel_ref": r["vessel_ref"], "space_status": r["space_status"],
        "sheet": r["sheet"], "source_row": r["source_row"],
        "carrier_scope": ",".join(r["carrier_scope"]) or None,
        "free_days_pol": r["free_days_pol"], "free_days_pod": r["free_days_pod"],
        "free_time_raw": r["free_time_raw"],
    } for r in d["routes"]]

    tarifas = [{
        "route_key": q["route_id"], "equipment_type": q["equipment_type"],
        "also_valid_for": ",".join(q["also_valid_for"]) or None,
        "base_rate": q["base_rate"], "currency": q["currency"], "unit": q["unit"],
        "adjusted_rate": q["adjusted_rate"], "weight_operator": q["weight_operator"],
        "weight_limit_ton": q["weight_limit_ton"], "weight_basis": q["weight_basis"],
        "cargo_type": q["cargo_type"], "rate_source": q["rate_source"],
        "raw_cell": q["raw_cell"],
    } for q in d["quotes"]]

    taxas = [dict(s, route_key=r["id"]) for r in d["routes"] for s in r["surcharges"]]

    # --- Parte A: staging + dimensões ---
    a = [cab, DDL_STAGE, "", "-- Dimensões (idempotentes)"]
    for lote in _valores(d["ports"], ["unlocode", "name", "country"], "ports"):
        a.append(lote.rstrip(";") + "\non conflict (unlocode) do update set name = excluded.name;")
    for lote in _valores(d["carriers"], ["code", "name"], "carriers"):
        a.append(lote.rstrip(";") + "\non conflict (code) do update set name = excluded.name;")
    partes = [("0004_freight_a_stage.sql", "\n".join(a) + "\n")]

    # --- Parte B: dados, fatiados por tamanho ---
    comandos = (
        _valores(rotas, ["route_key", "carrier", "pol", "pod", "trade_lane", "service_type",
                         "service_name", "validity_start", "validity_end", "validity_raw",
                         "vessel_ref", "space_status", "sheet", "source_row", "carrier_scope",
                         "free_days_pol", "free_days_pod", "free_time_raw"], "_stage_route")
        + _valores(tarifas, ["route_key", "equipment_type", "also_valid_for", "base_rate",
                             "currency", "unit", "adjusted_rate", "weight_operator",
                             "weight_limit_ton", "weight_basis", "cargo_type", "rate_source",
                             "raw_cell"], "_stage_rate")
        + _valores(taxas, ["route_key", "fee_code", "fee_label", "amount", "currency",
                           "charge_basis", "equipment_type", "min_weight_ton", "condition_raw",
                           "source_column"], "_stage_surcharge")
        + _valores(d["issues"], ["sheet", "source_row", "severity", "kind", "detail"],
                   "_stage_issue")
    )

    blocos, atual, tam = [], [], 0
    for c in comandos:
        # Nunca parte um INSERT ao meio: o corte só acontece entre comandos.
        if atual and tam + len(c) > limite_bytes:
            blocos.append(atual)
            atual, tam = [], 0
        atual.append(c)
        tam += len(c) + 1
    if atual:
        blocos.append(atual)

    for i, bloco in enumerate(blocos, 1):
        nome = f"0004_freight_b_data_{i:02d}.sql"
        topo = (f"{cab}-- PARTE B{i} de {len(blocos)} — dados para o staging.\n"
                f"-- Rode as partes B em ordem, depois a parte C.\n"
                "set search_path to mcat, public;\n")
        partes.append((nome, topo + "\n" + "\n".join(bloco) + "\n"))

    partes.append(("0004_freight_c_load.sql",
                   cab + _ddl_load(rs, len(rotas), len(tarifas))))
    return partes


def main():
    ap = argparse.ArgumentParser(description="Desagrupa uma rate sheet de frete marítimo.")
    ap.add_argument("arquivo", nargs="?", help="caminho do .xlsx")
    ap.add_argument("--year", type=int, default=dt.date.today().year,
                    help="ano das validades (a planilha não traz ano)")
    ap.add_argument("--json-out", default=os.path.join(RAIZ, "src", "data", "freightRates.json"))
    ap.add_argument("--sql-dir", default=os.path.join(RAIZ, "seeds", "freight"),
                    help="pasta das partes SQL (o SQL Editor do Supabase tem limite de tamanho)")
    ap.add_argument("--sql-chunk", type=int, default=350_000,
                    help="tamanho maximo de cada parte de dados, em bytes")
    a = ap.parse_args()

    caminho = a.arquivo
    if not caminho or not os.path.exists(caminho):
        cand = glob.glob(os.path.expanduser("~/Downloads/*rate*sheet*.xlsx"))
        if not cand:
            sys.exit("Planilha não encontrada. Informe o caminho do .xlsx.")
        caminho = cand[0]

    d = processar(caminho, a.year)

    os.makedirs(os.path.dirname(a.json_out), exist_ok=True)
    with open(a.json_out, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, separators=(",", ":"))
    os.makedirs(a.sql_dir, exist_ok=True)
    for antigo in os.listdir(a.sql_dir):
        if antigo.startswith("0004_freight_") and antigo.endswith(".sql"):
            os.remove(os.path.join(a.sql_dir, antigo))
    partes = gerar_partes(d, a.sql_chunk)
    for nome, conteudo in partes:
        with open(os.path.join(a.sql_dir, nome), "w", encoding="utf-8") as f:
            f.write(conteudo)

    sev = {}
    for i in d["issues"]:
        sev[i["severity"]] = sev.get(i["severity"], 0) + 1
    print(f"fonte      : {d['rate_sheet']['source_file']}")
    print(f"emissao    : {d['rate_sheet']['issued_on']} (ano assumido {a.year})")
    print(f"portos     : {len(d['ports'])}")
    print(f"armadores  : {len(d['carriers'])}")
    print(f"rotas      : {len(d['routes'])}")
    print(f"cotacoes   : {len(d['quotes'])}")
    print(f"issues     : {len(d['issues'])} {sev}")
    print(f"json -> {a.json_out}")
    print(f"sql  -> {a.sql_dir} ({len(partes)} partes)")
    for nome, conteudo in partes:
        print(f"        {nome:<32} {len(conteudo)/1024:7.0f} KB")


if __name__ == "__main__":
    main()
